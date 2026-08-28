import json, os, re, pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT, OUT = "/home/user/gurok-siparis", "/home/user/gurok-siparis/docs/bar-eksik"
SCR = "/tmp/claude-0/-home-user-gurok-siparis/86aebd71-8b4b-5444-8814-d27ada59b6ab/scratchpad"
os.makedirs(OUT, exist_ok=True)

D = json.loads(open(ROOT+"/veri.js", encoding="utf-8").read().split("\n")[3][len("const D="):].rstrip().rstrip(";"))
depolar = pickle.load(open(SCR+"/depolar.pkl","rb"))

# Tum barlarin urunlerinden kod -> uygulama grubu sozlugu (yeni kalemlere grup onerisi icin)
app_grup, app_ad = {}, {}
for o in D:
    for it in o["i"]:
        app_grup.setdefault(it["k"], it["g"]); app_ad.setdefault(it["k"], it["a"])

BIRIM = {"ad":"ADET","kol":"KOLI","fic":"FICI","bot":"SISE","kg":"KG","lt":"LT","pk":"PAKET","ktu":"KUTU"}
bar_by_kod = {o["c"]: o for o in D}

satirlar = []          # (outlet, ad, eksik[], fazla[])
for o in D:
    kod = o["c"]
    dep = depolar.get(kod)
    mevcut = {it["k"] for it in o["i"]}
    ln = dep["items"] if dep else {}
    eksik = sorted([v for k, v in ln.items() if k not in mevcut],
                   key=lambda x: (x["grup"] or "", x["a"]))
    fazla = sorted([it for it in o["i"] if it["k"] not in ln],
                   key=lambda x: (x["g"], x["a"]))
    satirlar.append((o, eksik, fazla))

# Uygulamada hic olmayan depolar
yok = {k: v for k, v in depolar.items() if k not in bar_by_kod}

wb = Workbook(); wb.remove(wb.active)
HDR = PatternFill("solid", fgColor="1F3864"); GRP = PatternFill("solid", fgColor="DCE6F1")
WARN = PatternFill("solid", fgColor="FCE4D6")
thin = Side(style="thin", color="BFBFBF"); BOX = Border(left=thin,right=thin,top=thin,bottom=thin)

ws0 = wb.create_sheet("00 OZET")
ws0.append(["KOD","BAR ADI","LISTEDE","LN'DE","EKSIK (LN'de var, listede yok)","FAZLA (listede var, LN'de hareket yok)"])
for o, eksik, fazla in satirlar:
    ws0.append([o["c"], o["n"], len(o["i"]), len(depolar.get(o["c"],{}).get("items",{})), len(eksik), len(fazla)])
for k, v in sorted(yok.items()):
    ws0.append([k, v["ad"] + "  (UYGULAMADA YOK)", 0, len(v["items"]), len(v["items"]), 0])
    for c in range(1,7): ws0.cell(ws0.max_row, c).fill = WARN
for c in range(1,7):
    h = ws0.cell(1,c); h.font = Font(bold=True, color="FFFFFF"); h.fill = HDR
    h.alignment = Alignment(wrap_text=True, horizontal="center")
for col,w in zip("ABCDEF",[10,32,10,10,18,20]): ws0.column_dimensions[col].width = w
ws0.freeze_panes = "A2"

def sayfa(kodetiket, baslik, eksik, fazla):
    ad = re.sub(r"[\\/*?:\[\]']", " ", baslik)[:24]
    ws = wb.create_sheet(f"{kodetiket} {ad}"[:31])
    ws.append([f"{kodetiket} - {baslik}"]); ws.cell(1,1).font = Font(bold=True, size=12)
    ws.append([f"LN'de hareket gormus ama BAR LISTESINDE OLMAYAN kalemler: {len(eksik)}"])
    ws.cell(2,1).font = Font(bold=True, color="C00000")
    ws.append(["EKLE","KOD","URUN ADI","BIRIM","LN KALEM GRUBU","ONERILEN APP GRUBU","CIKIS","KALAN"])
    for c in range(1,9):
        h = ws.cell(3,c); h.font = Font(bold=True, color="FFFFFF"); h.fill = HDR
        h.alignment = Alignment(wrap_text=True, horizontal="center")
    son = None
    for it in eksik:
        if it["grup"] != son:
            son = it["grup"]; ws.append(["","",f"— {son} —","","","","",""])
            for c in range(1,9):
                ws.cell(ws.max_row,c).fill = GRP; ws.cell(ws.max_row,c).font = Font(bold=True)
        ws.append(["", it["k"], it["a"], BIRIM.get(it["b"], it["b"].upper()),
                   it["grup"], app_grup.get(it["k"], "?  (yeni kalem)"), it["cikis"], it["kalan"]])
        for c in range(1,9): ws.cell(ws.max_row,c).border = BOX
    if fazla:
        ws.append([]); ws.append([f"BAR LISTESINDE VAR AMA LN'DE HAREKET YOK: {len(fazla)}"])
        ws.cell(ws.max_row,1).font = Font(bold=True, color="7F6000")
        ws.append(["CIKAR","KOD","URUN ADI","BIRIM","APP GRUBU","","",""])
        for c in range(1,9):
            h = ws.cell(ws.max_row,c); h.font = Font(bold=True, color="FFFFFF"); h.fill = HDR
        for it in fazla:
            ws.append(["", it["k"], it["a"], BIRIM.get(it["b"], it["b"].upper()), it["g"], "", "", ""])
            for c in range(1,9): ws.cell(ws.max_row,c).border = BOX
    for col,w in zip("ABCDEFGH",[7,14,44,8,34,26,10,10]): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

for o, eksik, fazla in satirlar:
    sayfa(o["c"][-3:], o["n"], eksik, fazla)
for k, v in sorted(yok.items()):
    sayfa(k[-3:], v["ad"], sorted(v["items"].values(), key=lambda x:(x["grup"] or "", x["a"])), [])

wb.save(OUT+"/BAR_EKSIK_URUNLER.xlsx")

# ---------- Markdown ----------
def md(kod, ad, eksik, fazla, listede):
    L = [f"# {kod} — {ad}", "",
         f"Bar listesinde: **{listede}** · LN raporunda: **{listede - len(fazla) + len(eksik)}**",
         "", f"## LN'de var, bar listesinde YOK — {len(eksik)} kalem", "",
         "Eklemek istediklerini `[x]` yap.", ""]
    son = None
    for it in eksik:
        if it["grup"] != son:
            son = it["grup"]; L += ["", f"### {son}", ""]
        g = app_grup.get(it["k"], "**YENI KALEM — grup gerekli**")
        L.append(f"- [ ] `{it['k']}` — {it['a']} ({BIRIM.get(it['b'], it['b'].upper())}) · grup: {g}")
    if fazla:
        L += ["", f"## Bar listesinde var, LN'de hareket yok — {len(fazla)} kalem", "",
              "Cikarmak istediklerini `[x]` yap.", ""]
        son = None
        for it in fazla:
            if it["g"] != son:
                son = it["g"]; L += ["", f"### {son}", ""]
            L.append(f"- [ ] `{it['k']}` — {it['a']} ({BIRIM.get(it['b'], it['b'].upper())})")
    fn = f"{kod}_{re.sub(r'[^A-Z0-9]+','_',ad.upper()).strip('_')}.md"
    open(OUT+"/"+fn, "w", encoding="utf-8").write("\n".join(L)+"\n")
    return fn

oz = ["# Bar Listeleri — LN Karsilastirma Raporu", "",
      "Kaynak: `KALEM DEPO ENVANTER HAREKETLERI (KUM)` — 18-08-26 tarihli iki PDF (26 depo).",
      "Karsilastirma: LN depo hareketleri  ⟷  `veri.js` bar listeleri.", "",
      "| Kod | Bar | Listede | LN'de | Eksik | Fazla | Detay |",
      "|---|---|---:|---:|---:|---:|---|"]
for o, eksik, fazla in satirlar:
    fn = md(o["c"], o["n"], eksik, fazla, len(o["i"]))
    oz.append(f"| {o['c']} | {o['n']} | {len(o['i'])} | {len(depolar.get(o['c'],{}).get('items',{}))} | **{len(eksik)}** | {len(fazla)} | [md]({fn}) |")
oz += ["", "## Uygulamada tanimli olmayan depolar", "", "| Kod | Depo | LN kalem | Detay |", "|---|---|---:|---|"]
for k, v in sorted(yok.items()):
    fn = md(k, v["ad"], sorted(v["items"].values(), key=lambda x:(x["grup"] or "", x["a"])), [], 0)
    oz.append(f"| {k} | {v['ad']} | {len(v['items'])} | [md]({fn}) |")
oz += ["", "## Nasil kullanilir", "",
       "1. `BAR_EKSIK_URUNLER.xlsx` — her bar ayri sayfa. Ust blok = **eklenecek adaylar** (`EKLE` sutununa `X`),",
       "   alt blok = **cikarilabilecekler** (`CIKAR` sutununa `X`).",
       "2. `ONERILEN APP GRUBU` sutunu, kalem baska bir barin listesinde varsa oradaki grubu gosterir;",
       "   `?  (yeni kalem)` yazanlar icin grup adini sen belirle.",
       "3. Isaretli dosyayi geri gonder, `veri.js` guncellensin.", ""]
open(OUT+"/00_OZET.md","w",encoding="utf-8").write("\n".join(oz)+"\n")

print(f"{'KOD':<8}{'BAR':<26}{'LISTE':>6}{'LN':>6}{'EKSIK':>7}{'FAZLA':>7}")
for o, eksik, fazla in satirlar:
    print(f"{o['c']:<8}{o['n']:<26}{len(o['i']):>6}{len(depolar.get(o['c'],{}).get('items',{})):>6}{len(eksik):>7}{len(fazla):>7}")
print("\nUygulamada olmayan depolar:")
for k,v in sorted(yok.items()): print(f"  {k}  {v['ad']:<28} {len(v['items'])} kalem")
