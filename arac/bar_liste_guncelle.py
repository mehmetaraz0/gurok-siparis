#!/usr/bin/env python3
"""Isaretlenmis BAR_EKSIK_URUNLER.xlsx dosyasindan veri.js bar listelerini gunceller.

Excel'de her depo icin bir sayfa vardir; kullanici eklemek ISTEMEDIGI satirlari
siler, kalan satirlar eklenecek kalemlerdir. Kalem adi/birimi LN PDF'inden
(depolar.pkl) alinir; boylece veri.js'teki ham birim kodlari (ad/kol/bot/...)
korunur ve Excel'deki gosterim adlarina bagimlilik olmaz.
"""
import json, pickle, sys, os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = sys.argv[1]
PKL  = sys.argv[2]

# Excel'in "ONERILEN APP GRUBU" sutunu bostaysa ("?" ile baslayan) kullanilacak
# yedek grup atamasi. Sutun doluysa DAIMA kullanicinin yazdigi grup gecerlidir.
YENI_GRUP = {
    "GNL07000035": "HIJYEN",                # KOLONYA DOKME
    "GNL07000203": "HIJYEN",                # ELDIVEN CERRAHI PUDRALI S
    "GNL18000045": "TEMIZLIK & DETERJAN",   # SUMA D10 5 KG (DEZENFEKTAN)
    "ICA02000031": "KOLA & GAZOZ",          # GAZOZ KUTU 200 ML 24 LU
    "ICA05000042": "CAYLAR",                # CAY MATCHA
    "ICB07000153": "VOTKALAR",              # VOTKA BELVEDERE PURE 70 CL
    "ICB07000216": "VISKILER",              # WHISKY MACALLAN 12 YRS 70 CL
    "ICB07000324": "CINLER",                # CIN LONDON NO:1 70 CL
    "ICB07000358": "TEKILALAR",             # TEQUILA EL JIMADOR 100 CL
    "YIY10000017": "YIYECEK",               # SEKER KUP DOKME
}

lines = open(f"{ROOT}/veri.js", encoding="utf-8").read().split("\n")
D = json.loads(lines[3][len("const D="):].rstrip().rstrip(";"))
depolar = pickle.load(open(PKL, "rb"))
GC = json.loads([l for l in lines if l.startswith("const GC=")][0][len("const GC="):].rstrip().rstrip(";"))

app_grup = {it["k"]: it["g"] for o in D for it in o["i"]}
# LN raporu kalem adlarini 25 karakterde kesiyor. Kod katalogda zaten varsa
# oradaki TAM ad/birim korunur; yoksa LN'in (kesik) adi kullanilir.
app_kalem = {it["k"]: it for o in D for it in o["i"]}
bar = {o["c"]: o for o in D}

# ---- Excel'den eklenecek kalemleri oku ----
wb = load_workbook(XLSX)
istek = {}   # depo kodu -> [kalem kodu]
for ws in wb.worksheets:
    if ws.title.startswith("00"):
        continue
    kod = "CSM" + ws.title.split()[0]
    kodlar = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and str(v).strip() and str(v).strip() != "KOD":
            g = (ws.cell(r, 6).value or "").strip()
            kodlar.append((str(v).strip(), "" if g.startswith("?") else g))
    if kodlar:
        istek[kod] = kodlar

# ---- Dogrulama: her kalem LN raporunda ve bir grupta olmali ----
hata = []
for depo, kodlar in istek.items():
    if depo not in depolar:
        hata.append(f"{depo}: LN raporunda boyle bir depo yok"); continue
    for k, g_xls in kodlar:
        if k not in depolar[depo]["items"]:
            hata.append(f"{depo}/{k}: LN raporunda bu depoda yok"); continue
        g = g_xls or app_grup.get(k) or YENI_GRUP.get(k)
        if not g:
            hata.append(f"{depo}/{k}: grup atanmamis ({depolar[depo]['items'][k]['a']})")
        elif g not in GC:
            hata.append(f"{depo}/{k}: '{g}' GC'de tanimli bir grup degil")
if hata:
    print("HATA:"); [print("  -", h) for h in hata]; sys.exit(1)

# ---- Uygula ----
eklenen, atlanan, yeni_outlet = {}, {}, []
for depo in sorted(istek):
    if depo not in bar:                      # uygulamada olmayan depo -> yeni outlet
        o = {"c": depo, "n": depolar[depo]["ad"], "i": []}
        D.append(o); bar[depo] = o; yeni_outlet.append(depo)
    o = bar[depo]
    mevcut = {it["k"] for it in o["i"]}
    n_ek = n_at = 0
    for k, g_xls in istek[depo]:
        if k in mevcut:
            n_at += 1; continue
        src = app_kalem.get(k) or depolar[depo]["items"][k]
        o["i"].append({"k": k, "a": src["a"], "b": src["b"],
                       "g": g_xls or app_grup.get(k) or YENI_GRUP[k]})
        mevcut.add(k); n_ek += 1
    eklenen[depo] = n_ek; atlanan[depo] = n_at

# Yeni outletler kod sirasina otursun
D.sort(key=lambda o: o["c"])

# GC'de karsiligi olmayan grup kalmadigini dogrula (aksi halde liste sonuna duser)
kayip = {it["g"] for o in D for it in o["i"]} - set(GC)
if kayip:
    print("UYARI: GC'de tanimsiz grup(lar):", kayip)

lines[3] = "const D=" + json.dumps(D, ensure_ascii=False, separators=(",", ":")) + ";"
open(f"{ROOT}/veri.js", "w", encoding="utf-8").write("\n".join(lines))

print(f"{'DEPO':<8}{'BAR':<28}{'EKLENEN':>8}{'ZATEN VAR':>10}{'YENI TOPLAM':>12}")
for depo in sorted(eklenen):
    if eklenen[depo] or atlanan[depo]:
        et = "  (YENI OUTLET)" if depo in yeni_outlet else ""
        print(f"{depo:<8}{bar[depo]['n']:<28}{eklenen[depo]:>8}{atlanan[depo]:>10}{len(bar[depo]['i']):>12}{et}")
print(f"\nToplam eklenen kalem: {sum(eklenen.values())}")
print(f"Yeni outlet: {', '.join(yeni_outlet) or 'yok'}")
print(f"Outlet sayisi: {len(D)}")
